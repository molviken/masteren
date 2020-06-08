import openpyxl
import os
from base64 import b64encode
from ttn import HandlerClient
from datetime import datetime
import paho.mqtt.client as mqtt

access_key = "ttn-account-v2.2aTBPpoyM78nhSgOKTkNE1At8OzeSUPxMCKlCbRC4jo"
app_id = "lora-nodes"
lora_dev1 = "lora_node1"
lora_dev2 = "lora_node2"
lora_dev3 = "lora_node3"
lte_node1 = "lte-node1"

Start_date = datetime.now()
curr_date = Start_date.strftime("%Y-%m-%d")
print(curr_date)
current_directory = os.getcwd()
main_node_file = current_directory+ "\\Logs\\lora_node2\\"+curr_date+".xlsx" # Node 2
wb_main = openpyxl.load_workbook(main_node_file)
sheet = wb_main.active
last_row = sheet.max_row
while sheet.cell(column=2, row=last_row).value is None and last_row > 0:
    last_row -= 1
last_col_b_value = sheet.cell(column=2, row=last_row).value
main_node_secs = int(last_col_b_value.strftime("%S"))


def sync_msg(node_name):
    global main_node_secs
    base_cmd = "11000000"
    final_directory = os.path.join(current_directory, (r''+'Logs'+'\\'+node_name))
    file_name = final_directory+'\\'+str(Start_date.date())+'.xlsx'
    wb  = openpyxl.load_workbook(file_name)
    sheet = wb.active
    last_row = sheet.max_row
    while sheet.cell(column=2, row=last_row).value is None and last_row > 0:
        last_row -= 1
    last_col_b_value = sheet.cell(column=2, row=last_row).value
    secs = int(last_col_b_value.strftime("%S"))

    if (abs(secs-main_node_secs) < 3):
        return "", False
    if (secs < main_node_secs):
        resultant = 60 - (main_node_secs-secs)
        print("less  ",resultant)
    if (secs > main_node_secs):
        resultant = secs - main_node_secs
        print("more  ",resultant)

    secs_hex = format(resultant, '02x')
    cmd = base_cmd + secs_hex
    cmd64 = b64encode(bytes(cmd, encoding="utf-8")).decode("utf-8")
    return cmd64, True

if __name__ == "__main__":
    handler = HandlerClient(app_id, access_key,discovery_address="discovery.thethings.network:1900")
    mqtt_application = handler.application()
    mqtt_client = handler.data()
    mqtt_client.connect()
    mqttc = mqtt.Client()
    mqttc.connect("mqtt.eclipse.org", 1883, 60)

    #msg_lte, sync_needed = sync_msg("lte-node1")
    #if (sync_needed):
     #   mqttc.publish("my/subscribe/topic",msg_lte)

    msg1, sync1 = sync_msg("lora_node1")
    if (sync1):
        mqtt_client.send(dev_id="lora_node1", pay=msg1, conf=False)
        

    msg3, sync3 = sync_msg("lora_node3")
    if (sync3):
        mqtt_client.send(dev_id="lora_node3", pay=msg3, conf=False)

    msg4, sync4 = sync_msg("my_board")
    if (sync4):
        mqtt_client.send(dev_id="my_board", pay=msg4, conf=False)



    mqttc.disconnect()
    mqtt_client.close()
    print("Ended")





