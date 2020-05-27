import os
from json import dumps
from datetime import datetime, timedelta
import csv
import dateutil.parser as dp
from math import pow
import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook
Datarates = ["SF12BW125", "SF11BW125", "SF10BW125","SF9BW125", "SF8BW125", "SF7BW125"]
current_directory = os.getcwd()

def payload_parser(payload):
    try:
        payload.decode("utf-8")
    except AttributeError:
        pass
    f_c = int(payload[0:2],16)
    bat = int(payload[2:4], 16)
    vbus = int(payload[4:8], 16)*0.004
    curr = int(payload[8:12], 16)*0.015
    return f_c, bat, vbus, curr

def timestamp_parser(ts):
    full_date_str = datetime.datetime.utcfromtimestamp(ts)
    date_str = full_date_str.strftime("%Y-%m-%d")
    clock_str = full_date_str.strftime("%H:%M")
    return date_str, clock_str

def WriteMetaToFile(payload, curr_date, dev_id):
    frame_count, bat_lvl, vbus, curr = payload_parser(payload)

    #Create directory for each day
    directory = 'Logs'+'\\'
    final_directory = os.path.join(current_directory, (r''+directory+dev_id))
    file_name = final_directory+"/"+str(curr_date.date())+".xlsx"
    if not os.path.exists(final_directory):
        os.makedirs(final_directory)
    try:
        wb = load_workbook(file_name)
        data_row = [frame_count, curr_date.time(), bat_lvl, vbus, curr, vbus*curr/1000]
        sheet = wb.active
        sheet.append(data_row)
        wb.save(file_name)
    
    except FileNotFoundError:
        header = ["Frame", "Time", "Battery", "Voltage [V]", "Current [mA]", "Power"]
        data_row = [frame_count, curr_date.time(), bat_lvl, vbus, curr, vbus*curr/1000]
        wb = Workbook()
        sheet = wb.active
        sheet.append(header)
        sheet.append(data_row)
        wb.save(file_name)


        
#today = datetime.datetime.now()
#payload_ex = "0435b17f355403210876"
#WriteMetaToFile(payload_ex, datetime.datetime.now() + datetime.timedelta(days=2), "lte-node2")