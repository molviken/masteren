import openpyxl
import os
from openpyxl.chart import LineChart,Reference 

current_dir = os.getcwd()
directory = current_dir+'\\'+'Logs'+'\\'
node = "lora_node2"+'\\'
day = "2020-05-24"
file_name = directory + node + day + ".xlsx"

wb = openpyxl.load_workbook(file_name)
sheet = wb.active 
last_row = sheet.max_row
while sheet.cell(column=2, row=last_row).value is None and last_row > 0:
        last_row -= 1

values = Reference(sheet, min_col = 3, min_row = 2, max_col = 3, max_row = last_row)

chart = LineChart()
chart.add_data(values)

chart.title = "LINE TEST"
chart.x_axis.title = "Time"
chart.y_axis.title = "Battery level"

sheet.add_chart(chart, "G2")

wb.save(file_name)